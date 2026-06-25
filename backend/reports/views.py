import csv
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework import permissions, status
from datetime import date, timedelta
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from tracking.models import DailyMealCompletion, WeightLog, WaterLog
from nutrition.models import MealPlan

class PDFReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Nutrition_Report_{user.username}.pdf"'

        # Setup document
        doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=15
        )
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#7f8c8d'),
            spaceAfter=25
        )
        h2_style = ParagraphStyle(
            'H2Style',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2980b9'),
            spaceBefore=15,
            spaceAfter=10
        )

        # Header
        story.append(Paragraph("Nutrition Management Platform", title_style))
        story.append(Paragraph(f"Personal Nutrition & Progress Report for {user.username} | Date: {date.today()}", subtitle_style))
        story.append(Spacer(1, 10))

        # Profile Stats
        profile = getattr(user, 'profile', None)
        if profile:
            profile_data = [
                ["Age", f"{profile.age or 'N/A'} yrs", "Weight", f"{profile.weight or 'N/A'} kg"],
                ["Height", f"{profile.height or 'N/A'} cm", "Gender", f"{profile.gender.capitalize()}"],
                ["Target Calories", f"{profile.daily_calorie_requirement} kcal", "Goal", f"{profile.goal.replace('_', ' ').capitalize()}"]
            ]
            t = Table(profile_data, colWidths=[120, 120, 120, 120])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f8f9fa')),
                ('TEXTCOLOR', (0,0), (-1,-1), colors.HexColor('#2c3e50')),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 8),
                ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#e2e8f0')),
            ]))
            story.append(Paragraph("Profile Details", h2_style))
            story.append(t)
            story.append(Spacer(1, 15))

        # Recent Logs (Last 7 Days)
        today = date.today()
        seven_days_ago = today - timedelta(days=7)
        completions = DailyMealCompletion.objects.filter(user=user, date__gte=seven_days_ago)
        
        log_data = [["Date", "Meal Type", "Calories", "Protein (g)", "Carbs (g)", "Fats (g)", "Status"]]
        for log in completions:
            status_txt = "Completed" if log.completed else "Logged"
            log_data.append([
                str(log.date),
                log.meal_type.capitalize(),
                f"{log.calories} kcal",
                f"{log.protein}g",
                f"{log.carbs}g",
                f"{log.fats}g",
                status_txt
            ])

        if len(log_data) > 1:
            t_log = Table(log_data, colWidths=[80, 80, 80, 80, 80, 80, 70])
            t_log.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#34495e')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f2f2f2')]),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#bdc3c7')),
            ]))
            story.append(Paragraph("Food Intake Log (Last 7 Days)", h2_style))
            story.append(t_log)
        else:
            story.append(Paragraph("No food items logged in the past week.", styles['Normal']))

        # Build PDF
        doc.build(story)
        return response


class ExcelReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Nutrition_Logs_{user.username}.xlsx"'

        wb = Workbook()
        
        # Sheet 1: Weight Logs
        ws_weight = wb.active
        ws_weight.title = "Weight History"
        ws_weight.append(["Date", "Weight (kg)"])
        for log in WeightLog.objects.filter(user=user).order_by('logged_at'):
            ws_weight.append([str(log.logged_at), log.weight])

        # Sheet 2: Water Logs
        ws_water = wb.create_sheet("Water History")
        ws_water.append(["Date", "Amount (ml)"])
        for log in WaterLog.objects.filter(user=user).order_by('date'):
            ws_water.append([str(log.date), log.amount])

        # Sheet 3: Meal Logs
        ws_meals = wb.create_sheet("Meal History")
        ws_meals.append(["Date", "Meal Type", "Calories (kcal)", "Protein (g)", "Carbs (g)", "Fats (g)", "Completed"])
        for log in DailyMealCompletion.objects.filter(user=user).order_by('-date'):
            ws_meals.append([str(log.date), log.meal_type, log.calories, log.protein, log.carbs, log.fats, log.completed])

        wb.save(response)
        return response


class CSVReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="Meal_History_{user.username}.csv"'

        writer = csv.writer(response)
        writer.writerow(["Date", "Meal Type", "Calories", "Protein", "Carbs", "Fats", "Completed"])
        
        logs = DailyMealCompletion.objects.filter(user=user).order_by('-date')
        for log in logs:
            writer.writerow([
                log.date, 
                log.meal_type, 
                log.calories, 
                log.protein, 
                log.carbs, 
                log.fats, 
                log.completed
            ])
            
        return response
